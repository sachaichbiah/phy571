from openpyxl import load_workbook
import imageio as io
import os
from PIL import Image
#file_names = sorted((fn for fn in os.listdir('.') if fn.startswith('surface')))
#making animation

wb = load_workbook(filename='Correlations64_20_1000_10_de 0.6 a 35.xlsx', read_only=True)
ws = wb['Sheet1']
N=ws.max_column
M=ws.max_row
Values=[]


def powerlaw_func(x,a,b,c): 

    return (a*(x**b)+c)

def exponenial_func(x, a, b, c):
    return a*np.exp(-b*x)+c


for i in range(0,N): 
    Data=[]
    for row in ws.rows:
    
        Data.append(row[i].value)
        
    Values.append(Data)



import matplotlib.pyplot as plt
from matplotlib import pylab

# Scientific libraries
import numpy as np
from scipy.optimize import curve_fit
 
NAME=[]
PoptE=[]
PoptP=[]
BETA=[]
X=np.linspace(0,len(Values[0])-3,len(Values[0])-2)
X[0]+=0.0001

p0e=(.7,.6,.2)
p0p=(0.85,-0.15,-0.35)

for i in range(1,34):#len(Values)): 
    Data=[]
    Beta=Values[i][0]
    for j in range(2,len(Values[0])):
        Data.append(Values[i][j])
    x=X
    y=Data
    
    
    #plt.figure()
    #plt.plot(X,Data)
    
    
    popte, pcove = curve_fit(exponenial_func, x, y, p0=p0e)
    poptp, pcovp = curve_fit(powerlaw_func,x,y, p0=p0p, maxfev=30000)
    xx = np.linspace(0, 30, 1000)
    yye = exponenial_func(xx, *popte)
    yyp = powerlaw_func(xx, *poptp)
    
    p0e=popte
    if Beta==2 or Beta==7 : 
        p0p=[-0.02490887 , 0.20527003 , 1.00482963]
    else : p0p = poptp
    
    plt.figure()
    plt.plot(x,y,'o', xx, yye, "green")
    plt.plot(xx,yyp, "red")
    pylab.title('Exponential (green) and Powerlaw (red) fits, Beta = '+str(Beta))
    #ax = plt.gca()
    #ax.set_axis_bgcolor((0.898, 0.898, 0.898))
    fig = plt.gcf()
    PoptE.append(1/popte[1])
    BETA.append(Beta)
    print(poptp, Beta)
    
    plt.savefig("Graphiques/CorrelationsForBeta"+str(Beta)+".png")
    NAME.append("Graphiques/CorrelationsForBeta"+str(Beta)+".png")
    
   
with io.get_writer('CurvesFit.gif', mode='I', duration=0.5) as writer:
    for filename in NAME:
        image = io.imread(filename)
        writer.append_data(image)
writer.close()

#plt.figure()
#plt.plot(BETA,PoptE)  
